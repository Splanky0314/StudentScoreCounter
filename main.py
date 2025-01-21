from openpyxl import load_workbook 
from typing import Union
from openpyxl.comments import Comment
import traceback

################################################
################################################
######### 매일 여기만 수정하면 됩니다 ###########

# 수정 필요 변수
TARGET_COL = 27 # 날짜 기반 1/20
search_row_range = 170
AUTHOR = "김다은"
EXCEL_FILE_NAME = "상벌점관리_01_20.xlsx" 
# EXCEL_FILE_NAME = "test.xlsx" 
KAKAO_TXT_NAME = "kakao11.txt" # 날짜 기반 1/20

###############################################
###############################################
###############################################
"""
필요 class 및 함수
"""

class ScoreData:
    def __init__(self, stdname: Union[str, list], score: int, detail: str):
        self.stdname = stdname # stdname 또는 stdname list 또는 roomnum이 저장
        self.score = score
        self.detail = detail

def excel_update_cell(ws, row:int, col:int, score_data: ScoreData):
    # update scores
    if ws.cell(row, col).value:
        ws.cell(row, col, score_data.score + ws.cell(row, col).value)
    else:
        ws.cell(row, col, score_data.score)

    # update comment
    if ws.cell(row, col).comment: 
        comment = Comment(ws.cell(row, col).comment.content + '\n' + score_data.stdname + ":" + score_data.detail + " " + str(score_data.score), AUTHOR)
    else: 
        comment = Comment(score_data.stdname + ":" + score_data.detail + " " + str(score_data.score), AUTHOR) #
    ws.cell(row, col).comment = comment


def find_command_and_return_index(line: str, command: str) -> ScoreData:
    words = list(line.split())
    idx = words.index(command)
    if command == 'ㅁㅁ':
        stdnamelist = list()
        score = int()
        detail = str()
        for i in range(idx+1, len(words)):
            if ('-' in words[i]) or ('+' in words[i]):
                score = int(words[i].replace('점', ''))
                detail = words[i+1]
                break
            else:
                stdnamelist.append(words[i])
        # return [stdnamelist, score, detail]
        return ScoreData(stdnamelist, score, detail)
    else: # ㄴㄴ, ㅇㅇ 처리
        stdname = words[idx+1]
        score = int(words[idx+2].replace('점', '').replace('호', ''))
        detail = words[idx+3]
        # return [stdname, score, detail]
        return ScoreData(stdname, score, detail)

###############################################
"""

student data 읽어오기
- excel 1행의 값을 stdname으로 인식
- excel 2행의 값을 roomnum으로 인식

"""
# excel 파일 열기
wb = load_workbook(EXCEL_FILE_NAME)
ws = wb.active

stdrow = dict()
room = dict()

# 학생명 - 엑셀 행번호 
for row in range(2, search_row_range):
    stdname = ws.cell(row, 1).value # excel: 1열에 stdname 저장되어 있어야
    if stdname: # 비어있는 셀은 건너 뜀
        stdrow[stdname] = row

print(stdrow)

# room number - 학생명 list
for row in range(2, search_row_range):
    stdname = ws.cell(row, 1).value # excel: 1열에 stdname 저장되어 있어야
    roomnum = ws.cell(row, 2).value # excel: 2열에 roomnum 저장되어 있어야

    # 비어있는 셀은 건너 뜀
    if not stdname or not roomnum:
        continue

    if roomnum not in room.keys():
        room[roomnum] = list()
    room[roomnum].append(stdname)

print(room)
###############################################

# kakao txt 읽어오기
f = open(KAKAO_TXT_NAME, 'rt', encoding='UTF8')

lines = f.readlines()
for line in lines: # kakao talk 내용 한줄 한줄 읽어오기
    try:
        if "ㅇㅇ" in line:
            """
            << 한명 상벌점(ㅇㅇ) 처리 >>
            ㅇㅇ 홍길동 +3 수학태도우수
            ㅇㅇ 홍길동 -3 수학태도불량
            """
            response = find_command_and_return_index(line, "ㅇㅇ")
            stdname = response.stdname
            score = response.score
            detail = response.detail      
            target_row = stdrow[stdname] # todo: 나중에 존재 X 경우 예외처리 

            # excel에 score & comment 업데이트
            excel_update_cell(ws, target_row, TARGET_COL, response)
            
            # console print
            print(target_row, TARGET_COL, stdname, score, detail)
                    

        elif "ㅁㅁ" in line:
            """
            << 여러명 상벌점(ㅁㅁ) 처리 >>
            ㅁㅁ 홍길동 김다은 김누구 +1 원어민zel상점
            ㅁㅁ 홍길동 김다은 김누구 -1 원어민zel벌점
            """
            
            response = find_command_and_return_index(line, "ㅁㅁ")
            stdnamelist = response.stdname
            score = response.score
            detail = response.detail      
            
            for stdname in stdnamelist:
                target_row = stdrow[stdname] # todo: 나중에 존재 X 경우 예외처리 

                # excel에 score & comment 업데이트
                excel_update_cell(ws, target_row, TARGET_COL, ScoreData(stdname, score, detail))
                
                # console print
                print(target_row, TARGET_COL, stdname, score, detail)
        

        elif "ㄴㄴ" in line:
            """
            << 룸 기준 상벌점(ㄴㄴ) 처리 >>
            ㄴㄴ 608 +3 청소우수
            """
            
            response = find_command_and_return_index(line, "ㄴㄴ")
            roomnum = int(response.stdname) # room number & '호' 제거
            score = response.score
            detail = response.detail     
            stdlist = room[roomnum] # 해당 room 학생들 리스트  

            for stdname in stdlist: 
                target_row = stdrow[stdname]

                # excel에 score & comment 업데이트
                excel_update_cell(ws, target_row, TARGET_COL, ScoreData(stdname, score, detail))
                
                # console print
                print(target_row, TARGET_COL, roomnum, stdname, score, detail)

    # Exception Handling
    except Exception as e: 
        print("="*30)
        print("에러 발생한 부분: " + line)
        traceback.print_exc() # print ERR Message 
        print("="*30)
        exit()

f.close()
wb.save(EXCEL_FILE_NAME)